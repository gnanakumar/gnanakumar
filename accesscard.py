from tkinter import *
import openpyxl  # A python module to handle Excel files
import datetime    # A python module to handle date and time
from tkinter import filedialog
import json

CONFIG_FILE_NAME = ".config.json"

# Get following parameters from GUI inputs.
params = {
    "receiver_email" : "user@email.com",
    "filename" : "",
    "column_Header" : ["Name", "Access card no", "Vehicle no","Valid till"],
    "filter_column":"Valid till",
    "notification_days" : 5,
    "window" :"600x150",
}

def update_configfile(file_name):
    params["receiver_email"]=svEmail.get()
    params["notification_days" ]=svNoOfDays.get()
    params["filename"]=svFileName.get()
    params["column_Header"]=svHeaders.get()
    params["filter_column"]=svFilterColumn.get()
    params["window"]=f'{root.winfo_width()}x{root.winfo_height()}+{root.winfo_x()}+{root.winfo_y()}'
    print(params["window"])
    with open(file_name, 'w') as fp:
        json.dump(params, fp)

def retrive_configfile(file_name):
    try:
        with open(file_name) as fp:
            cfg_params = json.load(fp)
            svEmail.set(cfg_params["receiver_email"])
            svNoOfDays.set(cfg_params["notification_days" ])
            svFileName.set(cfg_params["filename"])
            svHeaders.set(cfg_params["column_Header"])
            svFilterColumn.set(cfg_params["filter_column"])
            root.geometry(cfg_params["window"])
                        
    except IOError:
        svEmail.set(params["receiver_email"])
        svNoOfDays.set(params["notification_days" ])
        svFileName.set(params["filename"])
        svHeaders.set(params["column_Header"])
        svFilterColumn.set(params["filter_column"])
        root.geometry(params["window"])
        



def parse_xlsx_header(xlsx_file):
    '''
    Parse xlsx file and returen column number for expiry date and other output
    columns. Also returning data start row number
    '''
    data_locations = {
        "work_sheet" : None,
        "data_start_row": None,
        "columns": None,
    }
    # First interested column is expiry date

    data_locations["columns"] = [None]
    # Subsequent columns are requested bu user
    data_locations["columns"].extend([None for x in params["column_Header"]])
    num_locations = len(data_locations["columns"])

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    data_locations["work_sheet"] = ws

    for row in range(1, ws.max_row+1):
        for column in range(1, ws.max_column+1):
            if (ws.cell(row, column).value == params["filter_column"]):
                data_locations["columns"][0] = column
                data_locations["data_start_row"] = row + 1
                num_locations -= 1
            for column_label in params["column_Header"]:
                if ws.cell(row, column).value == column_label:
                    index = params["column_Header"].index(column_label)
                    data_locations["columns"][1+index] = column
                    num_locations -= 1
            if (num_locations == 0):
                break
        if (num_locations == 0):
            break

    return data_locations


def get_xlsx_data(xlsx_locations):
    '''
    Return expiry date and other output columns in a list from the excel file
    '''
    xlsx_data = []
    ws = xlsx_locations["work_sheet"]

    for row in range(xlsx_locations["data_start_row"], ws.max_row+1):
        row_data = []
        for column in xlsx_locations["columns"]:
            value = ws.cell(row, column).value
            row_data.append(value)
        if row_data[0] != None:
            xlsx_data.append(row_data)

    return xlsx_data


def filter_data(xlsx_data, num_expiry_days):
    '''
    Filter xlsx_data within num_expiry_days. Sort the filtered data 
    '''


def write_xlsx_file(xlsx_data):
    '''
    Write filtered data into output xlsx file
    '''
def browseFiles():
    filename = filedialog.askopenfilename(initialdir = ".",title = "Select a File",filetypes = (("Text files","*.xlsx*"),("all files","*.*")))
    svFileName.set(filename)


def Notify():
    lblStatusBar.config(text="Loading Headers..")
    data_loc= parse_xlsx_header(svFileName.get())
    lblStatusBar.config(text="Loading data..")
    data=get_xlsx_data(data_loc)
    lblStatusBar.config(text="Generate Report..")
    filter_data(data,svNoOfDays.get())
    write_xlsx_file(data)
    lblStatusBar.config(text="Finished Generate Report")
    
def on_closing():
    update_configfile(CONFIG_FILE_NAME)
    root.destroy()
    


if __name__ == '__main__':
    BG_COLOR="#158FAD"



    root = Tk()
    root.title('Access CARD Reminder')
    root.geometry("600x150")
    root.config(background = BG_COLOR)
    root.protocol("WM_DELETE_WINDOW", on_closing)

    svNoOfDays = StringVar()
    svEmail = StringVar()
    svFileName=StringVar()
    svHeaders = StringVar()
    svFilterColumn=StringVar()

    rootf = Frame(root,bg=BG_COLOR)
    rootf.pack()

    frame = Frame(root, padx=2, pady=2)
    frame.pack(side=BOTTOM)
    #filename
    lab = Label(rootf,text="Excel File Name:",bg=BG_COLOR)
    lab.grid(row="0",column="0" )

   
    enFile = Entry(rootf,textvariable = svFileName,width="60")

    enFile.grid(row="0",column="1",columnspan=3)


    button_explore = Button(rootf,text = "...",command =browseFiles)

    button_explore.grid(row="0",column="4")
    #email
    lab1 =Label(rootf,text="Email : ",bg=BG_COLOR)
    lab1.grid(row="1",column="0")

    enEmail = Entry(rootf,textvariable = svEmail,width="60")
    enEmail.grid(row="1",column="1",columnspan=3)

    #Column Headers
    lab3 =Label(rootf,text="Column headers",bg=BG_COLOR)
    lab3.grid(row="2",column="0")

    txtName3 = Entry(rootf,textvariable = svHeaders,width="60")
    txtName3.grid(row="2",column="1",columnspan=3)

    #Filter

    lab2 =Label(rootf,text="Filter Column: ",bg=BG_COLOR)
    lab2.grid(row="3",column="0")

    txtFil = Entry(rootf,textvariable = svFilterColumn,width="30")
    txtFil.grid(row="3",column="1")

    lab2 =Label(rootf,text="No of Days:",bg=BG_COLOR)
    lab2.grid(row="3",column="2")

    txtFil = Entry(rootf,textvariable = svNoOfDays,width="15")
    txtFil.grid(row="3",column="3")
 

    #buttons

    But = Button(rootf,text ="Send Notification ",command = Notify)
    But.grid(row="4",column="1")

    button_exit = Button(rootf,text = "Exit",command = on_closing,padx=10)
    button_exit.grid(row="4",column="2")

    #status bar


    lblTitle = Label(frame, text="Status: ")
    lblTitle.pack(side=LEFT)

    lblStatusBar = Label(frame, text="", width=300, bd=1, relief=SUNKEN)
    lblStatusBar.pack(side=RIGHT)
        
    retrive_configfile(CONFIG_FILE_NAME)

    root.mainloop()
